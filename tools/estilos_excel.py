from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import pandas as pd
from openpyxl.chart import LineChart, Reference, Series
# from tools.GeneradorGrafico import generarGrafico
from GeneradorGrafico import generarGrafico

colores = {
    "header_hora_fill": "203864",
    "header_fill": "4472C4",
    "header_font": "FFFFFF",
    "row_fill_1": "D9E1F2",
    "row_fill_2": "FFFFFF",
    "gen_actual_fill": "E2EFDA",
    "monto_fill": "FCE4D6",
    "consigna_fill": "FFF2CC",
    "totales_header_fill": "305496",
    "totales_header_font": "FFFFFF",
    "totales_gen_fill": "D9E1F2",
    "totales_data_font": "000000",
    # Colores adicionales (por si se requieren más variantes)
    "accent_1": "B4C7E7",
    "accent_2": "9BC2E6",
    "danger": "F8CBAD",
    "success": "E2EFDA",
}


def insertar_logo(worksheet, path_logo="assets/logo.png", logo_height_rows=4, logo_width_cols=5):
    try:
        img = Image(path_logo)
    
        col_width_pixels = 80
        img.width = logo_width_cols * col_width_pixels

        row_height_pts = 15 
        total_height_pts = logo_height_rows * row_height_pts * 1.8
        img.height = total_height_pts

        worksheet.add_image(img, 'A1')
        
        height_per_row = total_height_pts / logo_height_rows
        for i in range(1, logo_height_rows + 1):
            worksheet.row_dimensions[i].height = height_per_row
            
        return logo_height_rows
        
    except FileNotFoundError:
        print(f" Advertencia: No se encontró el logo en {path_logo}. Continuando sin logo.")
        return 0
    except Exception as e:
        print(f" Error al insertar la imagen: {e}.")
        return 0


def aplicar_formato_con_horas(writer, sheet_name, df):
    worksheet = writer.sheets[sheet_name]
    
    header_hora_fill = PatternFill(start_color=colores["header_hora_fill"], end_color=colores["header_hora_fill"], fill_type="solid")
    header_fill = PatternFill(start_color=colores["header_fill"], end_color=colores["header_fill"], fill_type="solid")
    header_font_hora = Font(bold=True, color=colores["header_font"], size=12)
    header_font = Font(bold=True, color=colores["header_font"], size=10)
    
    row_fill_1 = PatternFill(start_color=colores["row_fill_1"], end_color=colores["row_fill_1"], fill_type="solid")
    row_fill_2 = PatternFill(start_color=colores["row_fill_2"], end_color=colores["row_fill_2"], fill_type="solid")
    
    gen_actual_fill = PatternFill(start_color=colores["gen_actual_fill"], end_color=colores["gen_actual_fill"], fill_type="solid")
    monto_fill = PatternFill(start_color=colores["monto_fill"], end_color=colores["monto_fill"], fill_type="solid")
    consigna_fill = PatternFill(start_color=colores["consigna_fill"], end_color=colores["consigna_fill"], fill_type="solid")
    
    # Estilos para totales
    totales_header_fill = PatternFill(start_color=colores["totales_header_fill"], end_color=colores["totales_header_fill"], fill_type="solid")
    totales_header_font = Font(bold=True, color=colores["totales_header_font"], size=11)
    totales_gen_fill = PatternFill(start_color=colores["totales_gen_fill"], end_color=colores["totales_gen_fill"], fill_type="solid")
    totales_data_font = Font(bold=True, size=10, color=colores["totales_data_font"])
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    
    num_rows = worksheet.max_row
    num_cols = worksheet.max_column

    # print(f"Formateando hoja '{sheet_name}' con {num_rows} filas y {num_cols} columnas.")
    
    data_values = []
    column_names = []

    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        column_names.append(cell.value)
    
    for row in range(2, num_rows + 1):
        row_data = []
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            row_data.append(cell.value)
        data_values.append(row_data)
    
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None
    
    if worksheet.merged_cells:
        merged_ranges = list(worksheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            worksheet.unmerge_cells(str(merged_range))
    
    logo_offset_rows = insertar_logo(worksheet, path_logo="assets/logo.png", logo_height_rows=2, logo_width_cols=3)
    
    FILA_HORA_COMBINADA = 1 + logo_offset_rows 
    FILA_COLUMNAS = 2 + logo_offset_rows
    DATA_START_ROW = FILA_COLUMNAS + 1

    # print(f"Horas ordenadas encontradas: {df}")
    
    
    horas_ordenadas = df.attrs.get('horas_ordenadas', [])

    
    if horas_ordenadas:
        cell = worksheet.cell(row=FILA_HORA_COMBINADA, column=1)
        cell.value = 'FECHA'
        cell.fill = header_hora_fill
        cell.font = header_font_hora
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        
        cell = worksheet.cell(row=FILA_COLUMNAS, column=1)
        cell.value = 'FECHA'
        cell.fill = header_hora_fill
        cell.font = header_font_hora
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        
        worksheet.merge_cells(f'A{FILA_HORA_COMBINADA}:A{FILA_COLUMNAS}')
        
        cell = worksheet.cell(row=FILA_HORA_COMBINADA, column=2)
        cell.value = 'GENERADORA'
        cell.fill = header_hora_fill
        cell.font = header_font_hora
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        
        cell = worksheet.cell(row=FILA_COLUMNAS, column=2)
        cell.value = 'GENERADORA'
        cell.fill = header_hora_fill
        cell.font = header_font_hora
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        
        worksheet.merge_cells(f'B{FILA_HORA_COMBINADA}:B{FILA_COLUMNAS}')
        
        #por si quiero aumentar los datos por hora
        col_num = 3

        for hora in horas_ordenadas:
            start_col = col_num
            end_col = start_col + 2
            
            #transforma numero de columna a letra
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            
            cell = worksheet.cell(row=FILA_HORA_COMBINADA, column=start_col)
            cell.value = f'HORA {hora}'
            cell.fill = header_hora_fill
            cell.font = header_font_hora
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
            
            for c in range(start_col + 1, end_col + 1):
                cell = worksheet.cell(row=FILA_HORA_COMBINADA, column=c)
                cell.fill = header_hora_fill
                cell.border = thick_border
            
            worksheet.merge_cells(f'{start_letter}{FILA_HORA_COMBINADA}:{end_letter}{FILA_HORA_COMBINADA}')
            
            col_num += 3
    
    for col_num in range(1, len(column_names) + 1):
        if col_num > 2:
            cell = worksheet.cell(row=FILA_COLUMNAS, column=col_num)
            cell.value = column_names[col_num - 1]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        col_letter = get_column_letter(col_num)
        if col_num == 1:
            worksheet.column_dimensions[col_letter].width = 15 
        elif col_num == 2:
            worksheet.column_dimensions[col_letter].width = 25
        else:
            worksheet.column_dimensions[col_letter].width = 14
    
    for i, row_data in enumerate(data_values):
        row_num_excel = DATA_START_ROW + i
        row_fill = row_fill_1 if i % 2 == 0 else row_fill_2
        
        for col_num, value in enumerate(row_data, start=1):
            cell = worksheet.cell(row=row_num_excel, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            col_name = str(column_names[col_num - 1])
            
            if 'GEN.ACTUAL' in col_name:
                cell.fill = gen_actual_fill
            elif 'MONTO' in col_name:
                cell.fill = monto_fill
            elif 'CONSIGNA' in col_name:
                cell.fill = consigna_fill
                cell.font = Font(bold=True)
            elif col_num in [1, 2]:
                cell.fill = row_fill
            else:
                cell.fill = row_fill
            
            if col_num > 2 and isinstance(value, (int, float)):
                cell.number_format = '0'
            elif col_num == 1 and value:
                cell.number_format = 'YYYY-MM-DD'
    
    COLUMNA_INICIO_TOTALES = 5
    FILA_TOTALES_HEADER = DATA_START_ROW + len(data_values) + 2
    
    generadoras = df['GENERADORA'].unique()
    
    totales_por_gen = {}
    for gen in generadoras:
        df_gen = df[df['GENERADORA'] == gen]
        totales_por_gen[gen] = {}
        
        for col in df.columns:
            if col not in ['FECHA', 'GENERADORA'] and pd.api.types.is_numeric_dtype(df[col]):
                totales_por_gen[gen][col] = df_gen[col].sum()
    
  
    cell = worksheet.cell(row=FILA_TOTALES_HEADER, column=COLUMNA_INICIO_TOTALES)
    cell.value = "suma total"
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    
    cell = worksheet.cell(row=FILA_TOTALES_HEADER + 1, column=COLUMNA_INICIO_TOTALES)
    cell.value = "generadores"
    cell.fill = totales_gen_fill
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    
    col_totales_map = {}
    col_num = COLUMNA_INICIO_TOTALES + 1
    for col_name in column_names:
        if col_name not in ['FECHA', 'GENERADORA']:
            # Determinar el tipo de columna
            tipo = None
            if 'GEN.ACTUAL' in str(col_name):
                tipo = 'GEN.ACTUAL'
            elif 'MONTO' in str(col_name):
                tipo = 'MONTO'
            elif 'CONSIGNA' in str(col_name):
                tipo = 'CONSIGNA'
            
            if tipo and tipo not in col_totales_map:
                cell = worksheet.cell(row=FILA_TOTALES_HEADER, column=col_num)
                cell.value = tipo
                cell.fill = totales_header_fill
                cell.font = totales_header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
                col_totales_map[tipo] = col_num
                col_num += 1
    

    fila_actual = FILA_TOTALES_HEADER + 1
    
    for gen in generadoras:
        cell = worksheet.cell(row=fila_actual, column=COLUMNA_INICIO_TOTALES)
        cell.value = gen
        cell.fill = totales_gen_fill
        cell.font = Font(bold=False, size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        
        for tipo, col_num_total in col_totales_map.items():
            columnas_a_sumar = []
            for idx, col_name in enumerate(column_names, start=1):
                if tipo in str(col_name) and col_name not in ['FECHA', 'GENERADORA']:
                    columnas_a_sumar.append(get_column_letter(idx))
            
            if columnas_a_sumar:
                fila_datos = None
                for i, row_data in enumerate(data_values):
                    if row_data[1] == gen: 
                        fila_datos = DATA_START_ROW + i
                        break
                
                if fila_datos:
                    formula_partes = [f"{col}{fila_datos}" for col in columnas_a_sumar]
                    formula = f'=SUM({",".join(formula_partes)})'
                    
                    cell = worksheet.cell(row=fila_actual, column=col_num_total)
                    cell.value = formula
                else:
                    cell = worksheet.cell(row=fila_actual, column=col_num_total)
                    cell.value = 0
            else:
                cell = worksheet.cell(row=fila_actual, column=col_num_total)
                cell.value = 0
            
            # Aplicar color según tipo
            if tipo == 'GEN.ACTUAL':
                cell.fill = gen_actual_fill
            elif tipo == 'MONTO':
                cell.fill = monto_fill
            elif tipo == 'CONSIGNA':
                cell.fill = consigna_fill
            
            cell.font = totales_data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.number_format = '0'
        
        fila_actual += 1

    # Generar gráfico
    generarGrafico(column_names, data_values, generadoras,  worksheet, fila_actual , DATA_START_ROW)
