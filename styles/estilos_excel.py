from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def aplicar_formato_simple(writer, sheet_name, df):
    """Aplica formato simple sin encabezados combinados"""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Definir colores
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Colores alternos para filas
    row_fill_1 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Colores para diferentes tipos de datos
    gen_actual_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    monto_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    consigna_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formato de encabezados
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
        # Ajustar ancho de columna
        col_letter = get_column_letter(col_num)
        max_length = len(str(df.columns[col_num - 1]))
        worksheet.column_dimensions[col_letter].width = max(12, max_length + 2)
    
    # Formato de filas de datos
    for row_num in range(2, len(df) + 2):
        # Alternar color de filas
        row_fill = row_fill_1 if row_num % 2 == 0 else row_fill_2
        
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar color según el tipo de columna
            col_name = df.columns[col_num - 1]
            if 'GEN.ACTUAL' in str(col_name):
                cell.fill = gen_actual_fill
                cell.font = Font(bold=True)
            elif 'MONTO SUBE/BAJA' in str(col_name):
                cell.fill = monto_fill
            elif 'CONSIGNA' in str(col_name):
                cell.fill = consigna_fill
            elif 'FECHA' in str(col_name):
                cell.fill = row_fill
                # Aplicar formato de fecha
                if cell.value:
                    cell.number_format = 'YYYY-MM-DD'
            else:
                cell.fill = row_fill
            
            # Formato numérico para valores
            if col_num > 1 and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
    
    # Congelar primera fila y primera columna
    worksheet.freeze_panes = 'C2'
    
    # Ajustar altura de la primera fila
    worksheet.row_dimensions[1].height = 30

def aplicar_formato_con_horas(writer, sheet_name, df):
    """Aplica formato con encabezados de horas combinadas"""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Definir colores
    header_hora_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_hora = Font(bold=True, color="FFFFFF", size=12)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    
    # Colores alternos para filas
    row_fill_1 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Colores para diferentes tipos de datos
    gen_actual_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    monto_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    consigna_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    # Insertar una fila arriba para las horas
    worksheet.insert_rows(1)
    
    # Verificar si hay información de horas guardada
    if hasattr(df, 'attrs') and 'horas_ordenadas' in df.attrs:
        horas_ordenadas = df.attrs['horas_ordenadas']
        
        # Combinar celdas para "FECHA"
        worksheet.merge_cells('A1:A2')
        cell = worksheet['A1']
        cell.value = 'FECHA'
        cell.fill = header_hora_fill
        cell.font = header_font_hora
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        
        # Combinar celdas para "GENERADORA"
        worksheet.merge_cells('B1:B2')
        cell = worksheet['B1']
        cell.value = 'GENERADORA'
        cell.fill = header_hora_fill
        cell.font = header_font_hora
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        
        # Crear encabezados de horas combinadas
        col_num = 3  # Ahora empezamos en la columna C
        for hora in horas_ordenadas:
            start_col = col_num
            # Contar cuántas columnas tiene esta hora (debería ser 3)
            count = 3
            end_col = start_col + count - 1
            
            # Combinar celdas para la hora
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            worksheet.merge_cells(f'{start_letter}1:{end_letter}1')
            
            cell = worksheet[f'{start_letter}1']
            cell.value = f'HORA {hora}'
            cell.fill = header_hora_fill
            cell.font = header_font_hora
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
            
            col_num += count
    else:
        # Si no hay información de horas, solo combinar FECHA y GENERADORA
        worksheet.merge_cells('A1:A2')
        cell = worksheet['A1']
        cell.value = 'FECHA'
        cell.fill = header_hora_fill
        cell.font = header_font_hora
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        
        worksheet.merge_cells('B1:B2')
        cell = worksheet['B1']
        cell.value = 'GENERADORA'
        cell.fill = header_hora_fill
        cell.font = header_font_hora
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
    
    # Formato de encabezados de la segunda fila (nombres de columnas)
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=2, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
        # Ajustar ancho de columna
        col_letter = get_column_letter(col_num)
        if col_num == 1:
            worksheet.column_dimensions[col_letter].width = 15  # FECHA
        elif col_num == 2:
            worksheet.column_dimensions[col_letter].width = 20  # GENERADORA
        else:
            worksheet.column_dimensions[col_letter].width = 14
    
    # Formato de filas de datos (ahora comienzan en la fila 3)
    for row_num in range(3, len(df) + 3):
        row_fill = row_fill_1 if row_num % 2 == 1 else row_fill_2
        
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar color según el tipo de columna
            col_name = df.columns[col_num - 1]
            if 'Gen' in str(col_name):
                cell.fill = gen_actual_fill
                cell.font = Font(bold=True)
            elif 'Monto' in str(col_name):
                cell.fill = monto_fill
            elif 'Consigna' in str(col_name):
                cell.fill = consigna_fill
            else:
                cell.fill = row_fill
            
            # Formato numérico para valores
            if col_num > 2 and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
            # Formato de fecha para la primera columna (solo fecha, sin hora)
            elif col_num == 1 and cell.value:
                cell.number_format = 'YYYY-MM-DD'
    
    # Congelar primeras dos columnas y dos primeras filas
    worksheet.freeze_panes = 'C3'
    
    # Ajustar altura de las filas de encabezado
    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[2].height = 30